#!/usr/bin/env python
import os
import sys
import openpyxl


class UselessData(Exception):
    pass


def usage():
    print("""
    Usage: {} <source file> <destination file>
    """.format(os.path.basename(__file__)))
    sys.exit(1)

def adjust_column_widths(worksheet, padding='5'):
    """
    dynamically set the column width based on longest cell value
    """

    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column].width = length + 5
    return worksheet

def verify_row(row):
    """
    verifies that a row contains useful usefull data
    """
    # ignore cells that have 0 in source column K or L
    try:
        for cell in row:
            if cell.column in ['K','L'] and cell.value <= 0:
                raise UselessData()
    except UselessData:
        return False

    return True

if __name__ == '__main__':

    if not len(sys.argv) == 3:
        usage()

    SOURCE_COLUMNS = ['A','B','K','L','AM','BG','BU','BY']
    DEST_COLUMNS = ['A','B','C','D','E','F','G','H']

    source_wb = openpyxl.load_workbook(sys.argv[1])
    # assumes there is only one tab
    source_sheet = source_wb.active
    new_wb = openpyxl.Workbook()
    dest_sheet = new_wb.active


    row = 1
    for row_cells in source_sheet.iter_rows():

        if verify_row(row_cells):
            for cell in row_cells:
                if cell.column in SOURCE_COLUMNS:
                    # determine destination column
                    dest_col = DEST_COLUMNS[SOURCE_COLUMNS.index(cell.column)]
                    dest_cell = "{}{}".format(dest_col, row)

                    # save cell to new sheet
                    dest_sheet[dest_cell] = cell.value

                    # apply formula to row
                    formula = '=SUM((D{0}/C{0})-1)*-1'.format(row)

                    # create new column/cell with formula and formatting
                    new_col_cell = "I{}".format(row)
                    dest_sheet[new_col_cell] = formula
                    dest_sheet[new_col_cell].number_format = '0.00%'

            # move on to the next row
            row += 1

    # add header title for newly created column
    dest_sheet['I1'] = 'Discount'

    # apply some formatting
    dest_sheet = adjust_column_widths(dest_sheet)

    # do some more math
    dest_sheet['K1'] = "Average Discount"
    dest_sheet['K2'] = '=AVERAGE(I2:I{})'.format(row)

    # save the sheet
    print("Saving {} rows".format(row))
    new_wb.save(sys.argv[2])
